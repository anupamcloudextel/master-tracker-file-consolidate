"""
Read every .xlsx under Wireless/, take the "General Info" sheet (default header row 5),
and write one merged workbook under output/.

For "Small Cell Master Data Tracker" files, the header row is auto-detected by scanning
from row 1: the first row where any cell matches your mapping source column titles (or
distinctive text from them) is used as the header; otherwise UID / keywords / --header-row.
Use --no-auto-header to force --header-row only.

Base columns are taken as the maximum common header set across all successfully read files:
only headers present in every readable workbook are kept in the final output. Order follows
the first successfully read workbook.

If the output file already exists, it is removed first so each run produces a fresh file.

Use --only-files to consolidate specific workbooks by basename (searched under --input-dir).

Workbooks whose name contains "Small Cell Master Data Tracker" get column aliases applied
(source headers on the left of your mapping sheet -> canonical target headers on the right)
so their values line up with other files for consolidation.

The main output workbook includes a Base_headers sheet listing the common (intersection)
column names used for the matched export, for reference.
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
import traceback
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# (source_header, target_header) — left-side names from Small Cell Master Data trackers
# map into right-side / common template names. Multiline headers are normalized before match.
# Multiple source strings can map to the same target (template typos / variants).
MASTER_COLUMN_ALIASES: list[tuple[str, str]] = [
    ("First Operator (VIL / Airtel)", "Operator (VIL / Airtel)"),
    (
        "CE Site Address (As per Rent Agreement) without Location (Area) and Pincode",
        "Site Address",
    ),
    (
        "CE Site Address (As per Rent Agreement) without Location (Area and Pincode)",
        "Site Address",
    ),
    (
        "Billable Route Length to Customer (im mtr)",
        "Billable LMC Length to Customer (in mtr)",
    ),
    (
        "Billable LMC Length to Customer (im mtr)",
        "Billable LMC Length to Customer (in mtr)",
    ),
    ("CE Site Location (Area)", "City / Region"),
    (
        "Site Type (Camouflage / Pole)",
        "Site Type (Pole / WM with Cam / WM w/o Cam/ HPSC)",
    ),
    ("Camouflage Size /Pole Height", "Actual Pole Length (in Meters)"),
    ("Camouflage Size /Pole Hight", "Actual Pole Length (in Meters)"),
    (
        "Actual No. of Battery installed (Set of 4=1 No)",
        "Actual No. of Battery installed (Set of 4=1 No,If LFP then it is 1)",
    ),
    ("RFI Date", "RFI Date"),
]

_DUP_SUFFIX_RE = re.compile(r"__dup\d+$", re.IGNORECASE)


def strip_dup_suffix(col_name: str) -> str:
    return _DUP_SUFFIX_RE.sub("", str(col_name))


def should_apply_master_data_mapping(path: Path) -> bool:
    """Apply alias mapping for Small Cell Master Data Tracker workbooks."""
    return "small cell master data tracker" in path.name.casefold()


def find_columns_matching_source(df: pd.DataFrame, source_raw: str) -> list[str]:
    """Find actual column labels that match source_raw after header normalization."""
    want = normalize_header(source_raw)
    found: list[str] = []
    for c in df.columns:
        base = strip_dup_suffix(str(c))
        if normalize_header(base) == want:
            found.append(str(c))
    return found


def apply_master_data_column_mapping(path: Path, df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename / merge source columns into canonical target headers so rows align with
    newer templates and common-header consolidation.
    """
    if not should_apply_master_data_mapping(path):
        return df

    out = df.copy()
    # Group all source variants by normalized target name
    by_target: dict[str, list[str]] = defaultdict(list)
    for src_raw, tgt_raw in MASTER_COLUMN_ALIASES:
        tgt_norm = normalize_header(tgt_raw)
        if normalize_header(src_raw) == tgt_norm:
            continue  # no-op pair
        by_target[tgt_norm].append(src_raw)

    for tgt_norm, src_raw_list in by_target.items():
        source_col_labels: list[str] = []
        for src_raw in src_raw_list:
            source_col_labels.extend(find_columns_matching_source(out, src_raw))
        # unique, preserve order
        seen: set[str] = set()
        unique_sources = []
        for c in source_col_labels:
            if c not in seen:
                seen.add(c)
                unique_sources.append(c)

        if not unique_sources:
            continue

        merged_from_sources = out[unique_sources[0]]
        for c in unique_sources[1:]:
            merged_from_sources = merged_from_sources.combine_first(out[c])

        if tgt_norm in out.columns:
            # Prefer values already under target column; fill gaps from mapped sources
            combined = out[tgt_norm].combine_first(merged_from_sources)
            drop_cols = [c for c in unique_sources if c != tgt_norm]
            out = out.drop(columns=drop_cols, errors="ignore")
            out[tgt_norm] = combined
        else:
            out = out.drop(columns=unique_sources)
            out[tgt_norm] = merged_from_sources

    return out


def normalize_sheet_label(name: str) -> str:
    """Collapse whitespace so 'General  Info' / NBSP variants match 'General Info'."""
    s = str(name).replace("\u00a0", " ").replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(s.split()).strip().casefold()


def resolve_sheet_name(path: Path, requested: str) -> str:
    """
    Match the requested sheet name to an actual workbook sheet tab.

    Excel tabs can differ by invisible characters (NBSP), double spaces, or
    case; pandas/openpyxl require an exact string match for sheet_name=.
    """
    want = normalize_sheet_label(requested)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
    finally:
        wb.close()
    for actual in names:
        if normalize_sheet_label(actual) == want:
            return actual
    raise ValueError(
        f"No sheet matching {requested!r} (normalized). "
        f"Available sheets: {names!r}"
    )


def normalize_header(name: object) -> str:
    s = str(name).replace("\r\n", "\n").replace("\r", "\n")
    return " ".join(s.split()).strip()


# Lazily built from MASTER_COLUMN_ALIASES + UID (after normalize_header exists)
_marker_exact_set: frozenset[str] | None = None
_marker_substrings_cf: tuple[str, ...] = ()


def _ensure_mapping_header_markers_cache() -> None:
    """Exact normalized source headers + substring anchors for header-row detection."""
    global _marker_exact_set, _marker_substrings_cf
    if _marker_exact_set is not None:
        return
    exact: set[str] = set()
    for src, _ in MASTER_COLUMN_ALIASES:
        exact.add(normalize_header(src))
    exact.add(normalize_header("UID"))
    exact.add(normalize_header("Site UID"))
    _marker_exact_set = frozenset(exact)
    # If a cell contains any of these (casefold), treat row as header row (first match wins)
    _marker_substrings_cf = (
        "billable route length to customer",
        "first operator (vil / airtel)",
        "ce site address (as per rent agreement)",
        "ce site location (area)",
        "site type (camouflage / pole)",
        "camouflage size /pole",
        "actual no. of battery installed",
        "rfi date",
    )


def _row_matches_mapping_header_markers(row: object) -> bool:
    """True if this row looks like the General Info header row (mapping source columns)."""
    _ensure_mapping_header_markers_cache()
    assert _marker_exact_set is not None
    for v in row:
        nv = normalize_header(v)
        if not nv:
            continue
        if nv in _marker_exact_set:
            return True
        nfc = nv.casefold()
        for sub in _marker_substrings_cf:
            if sub in nfc:
                return True
    return False


def _header_row_non_empty_count(row: object) -> int:
    n = 0
    for v in row:
        if pd.isna(v):
            continue
        if isinstance(v, str) and not v.strip():
            continue
        n += 1
    return n


def _row_has_uid_header_cell(row: object) -> bool:
    for v in row:
        nv = normalize_header(v).casefold()
        if not nv:
            continue
        if nv in ("uid", "site uid"):
            return True
        # Some templates use "Site UID" with different spacing
        if nv.endswith(" uid") and len(nv) < 40:
            return True
    return False


def detect_header_row_1_indexed(
    path: Path,
    actual_sheet: str,
    prefer: int = 5,
    max_scan: int = 50,
) -> int:
    """
    Find the Excel row (1-based) that holds real table headers.

    For Small Cell Master Data Tracker files (auto header on): scan from the top and use
    the **first** row where any cell matches a known mapping source column title (from
    MASTER_COLUMN_ALIASES) or a distinctive substring of those headers. If none, fall
    back to a row containing UID / Site UID, then keyword scoring, then ``prefer``.
    """
    raw = pd.read_excel(
        path,
        sheet_name=actual_sheet,
        header=None,
        nrows=max_scan,
        engine="openpyxl",
        dtype=object,
    )
    if raw.empty:
        return prefer

    # 1) First row where any cell matches the manual mapping source headers (user sheets)
    for r in range(len(raw)):
        if _row_matches_mapping_header_markers(raw.iloc[r]):
            return r + 1

    uid_rows: list[int] = []
    for r in range(len(raw)):
        if _row_has_uid_header_cell(raw.iloc[r]):
            uid_rows.append(r)

    prefer0 = prefer - 1

    if uid_rows:
        if prefer0 in uid_rows:
            return prefer
        # Prefer UID row in typical header band with most filled header cells
        band = [r for r in uid_rows if 3 <= r <= 14]
        pool = band if band else uid_rows
        best = max(pool, key=lambda r: _header_row_non_empty_count(raw.iloc[r]))
        return best + 1

    # No UID cell: score rows by known header keywords (weak fallback)
    hints = (
        "rfi date",
        "operator",
        "site address",
        "city",
        "region",
        "pincode",
        "site type",
        "billable",
    )
    best_r, best_sc = prefer0, -1
    for r in range(min(len(raw), 25)):
        texts: list[str] = []
        for v in raw.iloc[r]:
            if pd.notna(v):
                texts.append(normalize_header(v).casefold())
        joined = " ".join(texts)
        sc = sum(1 for h in hints if h in joined)
        if sc > best_sc:
            best_sc, best_r = sc, r
    if best_sc >= 4:
        return best_r + 1
    return prefer


def read_general_info(
    path: Path,
    sheet_name: str,
    header_row_1_indexed: int,
    *,
    auto_detect_header: bool = False,
) -> tuple[pd.DataFrame, int]:
    actual_sheet = resolve_sheet_name(path, sheet_name)
    effective_1 = header_row_1_indexed
    if auto_detect_header:
        effective_1 = detect_header_row_1_indexed(
            path, actual_sheet, prefer=header_row_1_indexed
        )
    header_row_0_indexed = effective_1 - 1
    df = pd.read_excel(
        path,
        sheet_name=actual_sheet,
        header=header_row_0_indexed,
        engine="openpyxl",
    )
    return df, effective_1


def uniquify_duplicate_columns(columns: list[str]) -> list[str]:
    """Make column labels unique so pandas concat can align (keep order)."""
    counts: dict[str, int] = {}
    out: list[str] = []
    for c in columns:
        counts[c] = counts.get(c, 0) + 1
        if counts[c] == 1:
            out.append(c)
        else:
            out.append(f"{c}__dup{counts[c]}")
    return out


def normalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    normalized = [normalize_header(c) for c in out.columns]
    out.columns = uniquify_duplicate_columns(normalized)
    return out


SOURCE_COLUMN = "Source_File"


def insert_source_column(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    """
    Prepend Source_File with the workbook basename.

    If the sheet already has a column named Source_File (or same after header normalize),
    rename the sheet column so we do not create duplicate labels (which breaks concat/Excel).
    """
    out = df.copy()
    if SOURCE_COLUMN in out.columns:
        out = out.rename(columns={SOURCE_COLUMN: f"{SOURCE_COLUMN}__from_sheet"})
    out.insert(0, SOURCE_COLUMN, path.name)
    return out


def append_log(log_path: Path, line: str) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(line.rstrip() + "\n")


def iter_excel_files(folder: Path) -> list[Path]:
    if not folder.is_dir():
        raise NotADirectoryError(f"Not a directory: {folder}")
    # Include all Excel files under input-dir recursively.
    files = sorted(folder.rglob("*.xlsx"))
    # Skip Excel temporary lock files
    return [p for p in files if not p.name.startswith("~$")]


def filter_paths_by_basenames(
    paths: list[Path],
    only_basenames: list[str],
    input_dir: Path,
) -> list[Path]:
    """
    Keep only workbooks whose filename matches one of only_basenames (case-insensitive).
    Order follows only_basenames; first path wins if the same name appears in multiple folders.
    """
    by_cf: dict[str, list[Path]] = defaultdict(list)
    for p in paths:
        by_cf[p.name.casefold()].append(p)

    out: list[Path] = []
    missing: list[str] = []
    for raw in only_basenames:
        want = Path(raw).name
        key = want.casefold()
        if key in by_cf and by_cf[key]:
            out.append(by_cf[key].pop(0))
        else:
            missing.append(want)

    if missing:
        raise SystemExit(
            "These workbook names were not found under "
            f"{input_dir.resolve()} (recursive):\n  "
            + "\n  ".join(missing)
        )
    return out


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Merge 'General Info' from all Excel files in a folder into one workbook."
    )
    parser.add_argument(
        "--input-dir",
        default="Wireless",
        help="Folder containing input .xlsx files (default: Wireless).",
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Folder for the consolidated workbook (default: output).",
    )
    parser.add_argument(
        "--output-file",
        default="consolidated_general_info.xlsx",
        help="Output filename (default: consolidated_general_info.xlsx).",
    )
    parser.add_argument(
        "--matched-output-file",
        default="consolidated_general_info_common.xlsx",
        help="Output filename for common matched columns (default: consolidated_general_info_common.xlsx).",
    )
    parser.add_argument(
        "--sheet",
        default="General Info",
        help="Sheet name to read (default: General Info).",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=5,
        help="1-indexed Excel row used as column headers (default: 5).",
    )
    parser.add_argument(
        "--no-auto-header",
        action="store_true",
        help=(
            "Disable header-row auto-detection for Small Cell Master Data Tracker files; "
            "use --header-row only."
        ),
    )
    parser.add_argument(
        "--no-source-column",
        action="store_true",
        help="Do not add a Source_File column with the originating workbook name.",
    )
    parser.add_argument(
        "--only-files",
        nargs="+",
        default=None,
        metavar="NAME",
        help=(
            "Only consolidate these workbook file names (basename). "
            "Searched recursively under --input-dir. Case-insensitive match."
        ),
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / args.output_file
    matched_output_path = output_dir / args.matched_output_file
    log_path = output_dir / "consolidate_wireless.log.txt"
    # Start fresh every run.
    if log_path.exists():
        try:
            log_path.unlink()
        except PermissionError:
            # If it can't be deleted (rare), we'll just append.
            pass

    append_log(
        log_path,
        f"[{datetime.now().isoformat(timespec='seconds')}] Start run. "
        f"input_dir={input_dir.resolve()} sheet={args.sheet!r} header_row={args.header_row} "
        f"output_path={output_path.resolve()}",
    )

    paths = iter_excel_files(input_dir)
    if not paths:
        raise SystemExit(f"No .xlsx files found in {input_dir.resolve()}")

    if args.only_files:
        paths = filter_paths_by_basenames(paths, list(args.only_files), input_dir)
        append_log(
            log_path,
            f"[FILTER] --only-files: consolidating {len(paths)} workbook(s): "
            + ", ".join(p.name for p in paths),
        )

    loaded: list[tuple[Path, pd.DataFrame]] = []
    skipped: list[tuple[str, str]] = []
    audit_rows: list[dict[str, object]] = []

    for path in paths:
        try:
            rel = str(path.resolve().relative_to(input_dir.resolve()))
        except ValueError:
            rel = str(path)
        try:
            auto_hdr = should_apply_master_data_mapping(path) and not args.no_auto_header
            df, header_row_used = read_general_info(
                path,
                args.sheet,
                args.header_row,
                auto_detect_header=auto_hdr,
            )
            if auto_hdr and header_row_used != args.header_row:
                append_log(
                    log_path,
                    f"[HEADER] {path.name}: auto-detected header row {header_row_used} "
                    f"(default was {args.header_row})",
                )
            df = normalize_dataframe_columns(df)
            df = apply_master_data_column_mapping(path, df)
        except ValueError as e:
            append_log(
                log_path,
                f"[ERROR] {path.name}: ValueError while reading sheet {args.sheet!r} "
                f"(header_row={args.header_row}). Message={e}",
            )
            append_log(log_path, traceback.format_exc().rstrip())
            skipped.append((path.name, str(e)))
            audit_rows.append(
                {
                    "Relative_Path": rel,
                    "File_Name": path.name,
                    "Status": "skipped",
                    "Rows_Read": 0,
                    "Header_Row_Used": "",
                    "Error": str(e),
                }
            )
            continue
        except Exception as e:
            append_log(
                log_path,
                f"[ERROR] {path.name}: Exception while reading sheet {args.sheet!r} "
                f"(header_row={args.header_row}). Type={type(e).__name__} Message={e}",
            )
            append_log(log_path, traceback.format_exc().rstrip())
            skipped.append((path.name, str(e)))
            audit_rows.append(
                {
                    "Relative_Path": rel,
                    "File_Name": path.name,
                    "Status": "skipped",
                    "Rows_Read": 0,
                    "Header_Row_Used": "",
                    "Error": f"{type(e).__name__}: {e}",
                }
            )
            continue

        loaded.append((path, df))
        audit_rows.append(
            {
                "Relative_Path": rel,
                "File_Name": path.name,
                "Status": "loaded",
                "Rows_Read": len(df),
                "Header_Row_Used": header_row_used,
                "Error": "",
            }
        )

    if skipped:
        print("Skipped files:")
        for name, reason in skipped:
            print(f"  - {name}: {reason}")
            append_log(log_path, f"[SKIPPED] {name}: {reason}")

    if not loaded:
        raise SystemExit("No sheets could be read; nothing to write.")

    all_frames: list[pd.DataFrame] = []
    for path, df in loaded:
        all_df = df.copy()
        if not args.no_source_column:
            all_df = insert_source_column(all_df, path)
        all_frames.append(all_df)
    merged_all = pd.concat(all_frames, ignore_index=True, sort=False)

    first_columns = list(loaded[0][1].columns)
    common = set(first_columns)
    for _path, df in loaded[1:]:
        common &= set(df.columns)
    base_columns = [c for c in first_columns if c in common]

    if not base_columns:
        print(
            "WARNING: No column names are common across ALL loaded workbooks. "
            "Full consolidated file will still be written; the 'common columns' "
            "workbook will be skipped (nothing matches every file).",
            flush=True,
        )
        append_log(
            log_path,
            "[SCHEMA] No intersection of headers across all loaded files; "
            "matched-output file will not be written.",
        )
        merged = pd.DataFrame()
    else:
        append_log(
            log_path,
            f"[SCHEMA] Using {len(base_columns)} common base column(s) across "
            f"{len(loaded)} readable file(s).",
        )
        frames: list[pd.DataFrame] = []
        for path, df in loaded:
            df = df[base_columns].copy()
            if not args.no_source_column:
                df = insert_source_column(df, path)
            frames.append(df)

        merged = pd.concat(frames, ignore_index=True, sort=False)

    if output_path.is_file():
        try:
            output_path.unlink()
        except PermissionError:
            raise SystemExit(
                f"Cannot replace output file (close it in Excel if open): {output_path}"
            ) from None

    if base_columns and matched_output_path.is_file():
        try:
            matched_output_path.unlink()
        except PermissionError:
            raise SystemExit(
                f"Cannot replace matched output file (close it in Excel if open): {matched_output_path}"
            ) from None

    audit_df = pd.DataFrame(audit_rows)
    if not args.no_source_column and SOURCE_COLUMN in merged_all.columns:
        src_summary = (
            merged_all.groupby(SOURCE_COLUMN, dropna=False)
            .size()
            .reset_index(name="Row_Count")
            .sort_values(SOURCE_COLUMN)
        )
    else:
        src_summary = pd.DataFrame(columns=[SOURCE_COLUMN, "Row_Count"])

    if base_columns:
        schema_msg = (
            f"Headers present in every loaded workbook: {len(base_columns)} column(s). "
            f"See also {matched_output_path.name}."
        )
    else:
        schema_msg = (
            "No column name appears in every loaded workbook (intersection is empty). "
            "The full merge is still in sheet General Info; the common-columns file "
            "was not written. Check Consolidation_audit and align templates or mapping."
        )
    schema_note = pd.DataFrame({"Message": [schema_msg]})

    first_loaded_name = loaded[0][0].name
    ref_note = (
        f"Base headers = column names that appear in EVERY loaded workbook (intersection). "
        f"Column order follows the first loaded file: {first_loaded_name}. "
        f"Same set as sheet General Info in {matched_output_path.name} when that file is written."
    )
    if base_columns:
        base_headers_df = pd.DataFrame(
            {
                "Order": list(range(1, len(base_columns) + 1)),
                "Base_Header": base_columns,
                "Reference_note": [ref_note] + [""] * (len(base_columns) - 1),
            }
        )
    else:
        base_headers_df = pd.DataFrame(
            {
                "Order": [1],
                "Base_Header": [
                    "(No common base headers — no column name appears in every file.)"
                ],
                "Reference_note": [ref_note],
            }
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        merged_all.to_excel(writer, sheet_name="General Info", index=False)
        audit_df.to_excel(writer, sheet_name="Consolidation_audit", index=False)
        src_summary.to_excel(writer, sheet_name="Source_File_summary", index=False)
        schema_note.to_excel(writer, sheet_name="Schema_note", index=False)
        base_headers_df.to_excel(writer, sheet_name="Base_headers", index=False)

    if base_columns:
        with pd.ExcelWriter(matched_output_path, engine="openpyxl") as writer:
            merged.to_excel(writer, sheet_name="General Info", index=False)
    else:
        append_log(
            log_path,
            f"[SKIP] Matched columns file not written (no common headers): "
            f"{matched_output_path.resolve()}",
        )

    print(
        f"Wrote {len(merged_all)} rows from {len(loaded)} file(s) -> {output_path.resolve()}"
    )
    if not src_summary.empty:
        print(
            "Rows per Source_File (also in sheet 'Source_File_summary'): "
            + ", ".join(
                f"{r[SOURCE_COLUMN]}={int(r['Row_Count'])}"
                for _, r in src_summary.iterrows()
            )
        )
    if base_columns:
        print(
            f"Wrote {len(merged)} rows with {len(base_columns)} common matched column(s) "
            f"-> {matched_output_path.resolve()}"
        )
    else:
        print(
            f"Did not write common-columns file (no headers shared by every file): "
            f"{matched_output_path.name}"
        )
    append_log(
        log_path,
        f"[OK] Wrote full consolidated file -> {output_path.resolve()} "
        f"(rows={len(merged_all)} files={len(loaded)})",
    )
    if base_columns:
        append_log(
            log_path,
            f"[OK] Wrote common matched columns file -> {matched_output_path.resolve()} "
            f"(rows={len(merged)} common_columns={len(base_columns)})",
        )


if __name__ == "__main__":
    main()
